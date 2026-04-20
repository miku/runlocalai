#!/usr/bin/env -S uv run --script
#
# /// script
# dependencies = [
#   "openpyxl>=3.1",
#   "xlrd>=2.0",
#   "rich>=13",
# ]
# ///

import sys
from pathlib import Path

import openpyxl
import xlrd
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.rule import Rule

console = Console()


def find_header_row(rows, num_cols):
    """Heuristic: find the first row where a majority of cells are non-empty.

    Preamble rows typically have content only in the first column (merged cells),
    while the true header row has labels across most columns.
    """
    threshold = max(3, num_cols * 0.5)
    for idx, row in enumerate(rows):
        non_empty = sum(
            1 for c in row[:num_cols] if c is not None and str(c).strip() != ""
        )
        if non_empty >= threshold:
            return idx
    # Fallback: find the first row with at least 2 non-empty cells
    for idx, row in enumerate(rows):
        non_empty = sum(
            1 for c in row[:num_cols] if c is not None and str(c).strip() != ""
        )
        if non_empty >= 2:
            return idx
    return 0


def format_size(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.2f} MB"


def display_preamble(preamble_rows, num_cols):
    """Show preamble rows as flowing text (merged-cell summaries)."""
    console.print(Rule("[dim italic]Preamble[/]", style="dim"))
    for row in preamble_rows:
        parts = [
            str(c).strip()
            for c in row[:num_cols]
            if c is not None and str(c).strip() != ""
        ]
        if parts:
            console.print(f"  [dim]{' '.join(parts)}[/]")


def display_data_table(headers, data_rows, total_cols, total_data_rows):
    """Show the actual tabular data with proper column headers."""
    console.print(Rule("[dim italic]Data[/]", style="dim"))

    table = Table(show_header=True, header_style="bold magenta")
    for h in headers:
        table.add_column(h, max_width=30)

    for row in data_rows[:15]:
        table.add_row(
            *[str(c) if c is not None else "[dim]─[/]" for c in row[: len(headers)]]
        )

    console.print(table)

    if total_cols > len(headers):
        console.print(f"     [dim]… {total_cols - len(headers)} more column(s)[/]")
    if total_data_rows > 15:
        console.print(f"     [dim]… {total_data_rows - 15} more row(s)[/]")


def inspect_xlsx(filepath: Path) -> None:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    file_size = filepath.stat().st_size

    console.print(
        Panel(
            f"[bold]File:[/] {filepath}\n"
            f"[bold]Format:[/] XLSX (Office Open XML)\n"
            f"[bold]Size:[/] {format_size(file_size)}\n"
            f"[bold]Sheets:[/] {len(wb.sheetnames)}",
            title="📊 XLSX File Info",
            border_style="blue",
        )
    )

    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        console.print()
        console.print(f'  📋 [bold cyan]Sheet: "{name}"[/]')
        console.print(f"     Dimensions : {max_row} rows × {max_col} cols")

        if max_row == 0 or max_col == 0:
            console.print("     [dim](empty sheet)[/]")
            continue

        all_rows = list(ws.iter_rows(min_row=1, max_row=max_row, values_only=True))

        header_idx = find_header_row(all_rows, max_col)

        # --- Preamble ---
        if header_idx > 0:
            display_preamble(all_rows[:header_idx], max_col)

        # --- Header ---
        raw_header = all_rows[header_idx]
        display_cols = min(max_col, 10)
        headers = [
            str(raw_header[i]).strip()
            if raw_header[i] is not None and str(raw_header[i]).strip()
            else f"Col {i + 1}"
            for i in range(display_cols)
        ]

        # --- Data ---
        data_rows = all_rows[header_idx + 1 :]
        # Skip trailing empty rows
        while data_rows and all(
            c is None or str(c).strip() == "" for c in data_rows[-1]
        ):
            data_rows.pop()

        display_data_table(headers, data_rows, max_col, len(data_rows))

    wb.close()


def inspect_xls(filepath: Path) -> None:
    wb = xlrd.open_workbook(filepath)

    file_size = filepath.stat().st_size

    console.print(
        Panel(
            f"[bold]File:[/] {filepath}\n"
            f"[bold]Format:[/] XLS (BIFF{wb.biff_version})\n"
            f"[bold]Size:[/] {format_size(file_size)}\n"
            f"[bold]Sheets:[/] {wb.nsheets}",
            title="📊 XLS File Info",
            border_style="green",
        )
    )

    for idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(idx)

        console.print()
        console.print(f'  📋 [bold cyan]Sheet: "{sheet.name}"[/]')
        console.print(f"     Dimensions : {sheet.nrows} rows × {sheet.ncols} cols")

        if sheet.nrows == 0 or sheet.ncols == 0:
            console.print("     [dim](empty sheet)[/]")
            continue

        all_rows = [
            [sheet.cell(r, c).value for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]

        header_idx = find_header_row(all_rows, sheet.ncols)

        if header_idx > 0:
            display_preamble(all_rows[:header_idx], sheet.ncols)

        raw_header = all_rows[header_idx]
        display_cols = min(sheet.ncols, 10)
        headers = [
            str(raw_header[i]).strip()
            if raw_header[i] is not None and str(raw_header[i]).strip()
            else f"Col {i + 1}"
            for i in range(display_cols)
        ]

        data_rows = all_rows[header_idx + 1 :]
        while data_rows and all(
            c is None or str(c).strip() == "" for c in data_rows[-1]
        ):
            data_rows.pop()

        display_data_table(headers, data_rows, sheet.ncols, len(data_rows))


def main():
    if len(sys.argv) < 2:
        console.print("[bold red]Usage:[/] uv run inspect_xl.py <file.xls|xlsx>")
        console.print(
            "Displays information and a preview of XLS/XLSX spreadsheet files."
        )
        sys.exit(1)

    filepath = Path(sys.argv[1])

    if not filepath.exists():
        console.print(f"[bold red]Error:[/] File not found: {filepath}")
        sys.exit(1)

    suffix = filepath.suffix.lower()

    if suffix == ".xlsx":
        inspect_xlsx(filepath)
    elif suffix == ".xls":
        inspect_xls(filepath)
    else:
        console.print(
            f"[bold red]Error:[/] Unsupported format '{suffix}'. Use .xls or .xlsx files."
        )
        sys.exit(1)


if __name__ == "__main__":
    main()
